from typing import ContextManager
from django.contrib import messages
from django.db.models import query
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse

from products.models import Product, MainCategory, Category, SubCategory
from accounts.models import State, UserData, CustomUser, PremiumUser

from django.core.mail import EmailMessage
from django.conf import settings

from django.contrib.auth.decorators import login_required
from .decorators import allowed_group

from .thread import ProductEnquiryThread, EnquiryThread


# def add_product_database():
#     path = (r'C:\Users\PARTH\Desktop\Grofusion\static\main\xlsx\grofusion.xlsx')
#     workbook = openpyxl.load_workbook(path)

    # worksheet = workbook['main category']
    # col = worksheet.max_row
    # for i in range(1, col+1):
    #     value = worksheet.cell(row=i, column=1).value
    #     name = worksheet.cell(row=i, column=2).value
    #     name = name.replace(' ', '')
    #     name = name.replace(',', '')
    #     name = name.replace('&', '')
    #     name = name.replace('-', '')
    #     name = name.replace('.', '')
    #     name = name.replace("'", '')
    #     name = name.replace('/', '')
    #     name = name.lower()
    #     MainCategory(name=name, value=value).save()
    #     print(i)
    
    # worksheet = workbook['category']
    # col = worksheet.max_row
    # for i in range(1, col+1):
    #     value = worksheet.cell(row=i, column=1).value
    #     name = worksheet.cell(row=i, column=2).value
    #     name = name.replace(' ', '')
    #     name = name.replace(',', '')
    #     name = name.replace('&', '')
    #     name = name.replace('-', '')
    #     name = name.replace('.', '')
    #     name = name.replace("'", '')
    #     name = name.replace('/', '')
    #     name = name.lower()
    #     Category(name=name, value=value).save()
    #     print(i)
    
    # worksheet = workbook['sub category']
    # col = worksheet.max_row
    # for i in range(1, col+1):
    #     value = worksheet.cell(row=i, column=1).value
    #     name = worksheet.cell(row=i, column=2).value
    #     name = name.replace(' ', '')
    #     name = name.replace(',', '')
    #     name = name.replace('&', '')
    #     name = name.replace('-', '')
    #     name = name.replace('.', '')
    #     name = name.replace("'", '')
    #     name = name.replace('/', '')
    #     name = name.lower()
    #     SubCategory(name=name, value=value).save()
    #     print(i)
    
    # worksheet = workbook['states']
    # col = worksheet.max_row
    # for i in range(1, col+1):
    #     value = worksheet.cell(row=i, column=1).value
    #     name = worksheet.cell(row=i, column=2).value
    #     State(name=name, value=value).save()
    #     print(i)




# Create your views here.

def index(request):
    # add_product_database()
    # Category.objects.all().delete()
    # SubCategory.objects.all().delete()
    products = Product.objects.filter(status=True).order_by('-id')
    products_agriculture = Product.objects.filter(main_category = MainCategory.objects.get(value='28'), status=True)
    products_building = Product.objects.filter(main_category = MainCategory.objects.get(value='25'), status=True)
    products_electronics = Product.objects.filter(main_category = MainCategory.objects.get(value='1'), status=True)
    products_beauty = Product.objects.filter(main_category = MainCategory.objects.get(value='14'), status=True)
    products_fashion = Product.objects.filter(main_category = MainCategory.objects.get(value='5'), status=True)
    products_homelight = Product.objects.filter(main_category = MainCategory.objects.get(value='9'), status=True)
    products_automobiles = Product.objects.filter(main_category = MainCategory.objects.get(value='26'), status=True)
    products_jewellery = Product.objects.filter(main_category = MainCategory.objects.get(value='8'), status=True)
    
    context = {'products': products, 'products_agriculture':products_agriculture, 'products_building':products_building, 'products_electronics':products_electronics, 'products_beauty':products_beauty, 'products_fashion':products_fashion, 'products_homelight':products_homelight, 'products_automobiles':products_automobiles, 'products_jewellery':products_jewellery}
    
    return render(request, 'main/index.html', context)


@login_required(login_url='login')
def dashboard(request):
    if request.user.type == 'Seller':
        return render(request, 'main/dashboard.html')

    elif request.user.type == 'Buyer':
        context = {}
        user = CustomUser.objects.get(email=request.user.email)
        if UserData.objects.filter(user=user).exists():
                data = UserData.objects.get(user=user)
                context = {'address': data.address, 'state': data.state.value, 'city': data.city, 'pincode': data.pincode}

        return render(request, 'main/dashboard-buyer.html', context)

    return redirect('index')

@login_required(login_url='login')
@allowed_group('Seller')
def product_listing(request):
    context = {}
    if Product.objects.filter(supplier=request.user).exists():
        products = Product.objects.filter(supplier=request.user).order_by('-id')
        context = {'products': products}
        
    return render(request, 'main/product-listing.html', context)

@login_required(login_url='login')
@allowed_group('Seller')
def add_product(request):
    return render(request, 'main/add-product.html')

# @login_required(login_url='login')
# def dashboard_buyer(request):

#     context = {}
#     user = CustomUser.objects.get(email=request.user.email)
#     if UserData.objects.filter(user=user).exists():
#             data = UserData.objects.get(user=user)
#             context = {'address': data.address, 'state': data.state.value, 'city': data.city, 'pincode': data.pincode}

#     return render(request, 'main/dashboard-buyer.html', context)

def crunch(name):
    name = name.replace(' ', '')
    name = name.replace(',', '')
    name = name.replace('&', '')
    name = name.replace('-', '')
    name = name.replace('.', '')
    name = name.replace("'", '')
    name = name.replace('/', '')
    name = name.lower()

    return name

def category_products(request, main_category=None, category=None, sub_category=None):
    # print(category)
    # print(sub_category)
    # print(main_category)

    products = None
    name = ''
    if sub_category != None:
        name = sub_category.replace('-', ' ').title()
        main_category = crunch(main_category)
        category = crunch(category)
        sub_category = crunch(sub_category)
        mc = get_object_or_404(MainCategory, name=main_category)
        c = get_object_or_404(Category, name=category)
        sc = get_object_or_404(SubCategory, name=sub_category)
        products = Product.objects.filter(main_category=mc, category=c, sub_category=sc, status=True)

    elif category != None:
        name = category.replace('-', ' ').title()
        main_category = crunch(main_category)
        category = crunch(category)
        mc = get_object_or_404(MainCategory, name=main_category)
        c = get_object_or_404(Category, name=category)
        # print(mc.name, mc.value)
        # print(c.name, c.value)
        products = Product.objects.filter(main_category=mc, category=c, status=True)

    else:
        name = main_category.replace('-', ' ').title()
        main_category = crunch(main_category)
        mc = get_object_or_404(MainCategory, name=main_category)
        products = Product.objects.filter(main_category=mc, status=True)

    total = len(products)
    context = {'products': products, 'category': name, 'total': total}
    return render(request, 'main/products-by-category.html', context)


def search(request):
    keyword = request.GET['query']
    if len(keyword) > 80:
        results = Product.objects.none()

    else:
        results_name = Product.objects.filter(product_name__icontains=keyword, status=True)
        results_main_category = Product.objects.filter(main_category__name__icontains=keyword, status=True)
        results_category = Product.objects.filter(category__name__icontains=keyword, status=True)
        results_sub_category = Product.objects.filter(sub_category__name__icontains=keyword, status=True)
        results = results_name.union(results_main_category, results_category, results_sub_category)

    total = len(results)
    context = {'results': results, 'total': total, 'query': keyword}

    return render(request, 'main/search-results.html', context)


def product_enquiry(request):
    if request.method == 'POST':
        name = request.POST['ffname']
        phone = request.POST['ffphone']
        email = request.POST['ffemail']
        product_id = request.POST['product_id']
        quantity = request.POST['qty']

        product = Product.objects.get(id=product_id)
        product_name = product.product_name
        user = product.supplier
        user_name = user.name

        recipient = ['guptaparth716@gmail.com']

        if PremiumUser.objects.filter(user=user).exists():
            recipient.append(user.email)

        ProductEnquiryThread(name, phone, email, product_name, product_id, user_name, quantity, recipient).start()
        # print(recipient)
        messages.success(request, 'Your enquiry has been sent, we will respond to you shortly...')

    return redirect('index')

def for_buyers(request):
    return render(request, 'main/for-buyers.html')

def for_sellers(request):
    return render(request, 'main/for-suppliers.html')

def about(request):
    return render(request, 'main/about.html')

def contact(request):

    if request.method == "POST":
        name = request.POST['name']
        email = request.POST['email']
        subject = request.POST['subject']
        message = request.POST['message']

        recipient = ['guptaparth716@gmail.com']

        EnquiryThread(name, email, subject, message, recipient).start()
        messages.success(request, 'Your enquiry has been sent, we will respond to you shortly...')

        return redirect('index')

    return render(request, 'main/contact.html')

def affiliate(request):
    return render(request, 'main/affiliate-program.html')

def all_categories(request):
    return render(request, 'main/all-categories.html')